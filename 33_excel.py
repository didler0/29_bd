import pyodbc
from openpyxl import Workbook
import db_adapter
from openpyxl import load_workbook

# Create a new workbook
workbook = Workbook()

# Export data from the Clients table
list_of_clients = db_adapter.SQL.GetClients()
worksheet_clients = workbook.create_sheet("Clients")
headers = ["Id", "PhoneNumber", "FirstName", "LastName", "MiddleName", "Address"]
worksheet_clients.append(headers)

for row in list_of_clients:
    row_list = list(row)
    worksheet_clients.append(row_list)

# Export data from the Hotels table
list_of_hotels = db_adapter.SQL.GetHotel()
worksheet_hotels = workbook.create_sheet("Hotels")
headers = ["Id", "HotelPhoneNumber", "Country", "Address"]
worksheet_hotels.append(headers)

for row in list_of_hotels:
    row_list = list(row)
    worksheet_hotels.append(row_list)

# Export data from the Routes table
list_of_routes = db_adapter.SQL.GetRoute()
worksheet_routes = workbook.create_sheet("Routes")
headers = ["Id", "HotelClass", "HotelId", "ClimateType"]
worksheet_routes.append(headers)

for row in list_of_routes:
    row_list = list(row)
    worksheet_routes.append(row_list)

# Export data from the Discounts table
list_of_discounts = db_adapter.SQL.GetDiscount()
worksheet_discounts = workbook.create_sheet("Discounts")
headers = ["Id", "DiscountPercentage", "NumberOfTours"]
worksheet_discounts.append(headers)

for row in list_of_discounts:
    row_list = list(row)
    worksheet_discounts.append(row_list)

# Export data from the Trips table
list_of_trips = db_adapter.SQL.GetTrip()
worksheet_trips = workbook.create_sheet("Trips")
headers = ["Id", "ClientId", "RouteId", "DepartureDate", "ArrivalDate", "DiscountId", "NumberOfDays"]
worksheet_trips.append(headers)

for row in list_of_trips:
    row_list = list(row)
    worksheet_trips.append(row_list)




# Save the workbook
workbook.save("BAZA.xlsx")

# Load the Excel file
workbook = load_workbook("BAZA.xlsx")

# Get the default sheet ("Sheet" in this case)
sheet_to_remove = workbook["Sheet"]

# Remove the sheet
workbook.remove(sheet_to_remove)

# Save the changes to the Excel file
workbook.save("BAZA.xlsx")
